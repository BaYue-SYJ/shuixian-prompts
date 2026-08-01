import openpyxl
import os
import random

base = r"C:\Users\lianxiang\Desktop\split_by_style\split_by_style\6大类"
files = sorted(os.listdir(base))

for fname in files:
    if not fname.endswith('.xlsx'):
        continue
    fpath = os.path.join(base, fname)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        
        # Collect all titles
        all_titles = []
        all_prompts = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True):
            if row[0]:
                all_titles.append(str(row[0]))
            if row[1]:
                all_prompts.append(str(row[1]))
        
        total = len(all_titles)
        
        # Sample random prompts from different parts of the file
        if all_prompts:
            # Pick samples from beginning, middle, and end
            sample_indices = [0]
            if total > 10:
                sample_indices.append(total // 4)
            if total > 20:
                sample_indices.append(total // 2)
            if total > 40:
                sample_indices.append(3 * total // 4)
            if total > 60:
                sample_indices.append(total - 5)
            
            print(f"Total entries: {total}")
            print(f"\n--- Sample titles from different sections ---")
            seen = set()
            for idx in sample_indices:
                if idx < len(all_titles) and all_titles[idx] not in seen:
                    seen.add(all_titles[idx])
                    print(f"  [{idx:4d}] {all_titles[idx]}")
            
            # Count some keyword patterns
            keywords = {}
            for t in all_titles:
                for kw in ['海报', '广告', 'UI', '信息图', '肖像', '游戏', '电影', '动漫', '摄影',
                           '3D', '科幻', '时尚', '杂志', '产品', '落地页', 'Slides', 'PPT',
                           '漫画', '插画', '水墨', '国风', '证件照', '网红', '直播', '品牌',
                           'logo', '设计', '概念', '壁纸', '人像', '场景', '卡牌', '分镜']:
                    if kw.lower() in t.lower():
                        keywords[kw] = keywords.get(kw, 0) + 1
            
            print(f"\n--- Keyword frequency in titles ---")
            for kw, cnt in sorted(keywords.items(), key=lambda x: -x[1])[:25]:
                print(f"  {kw}: {cnt}")
        
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
