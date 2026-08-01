"""
Extract all prompts + images from 6 Excel files.
Output: style-gallery/data/[category].json + images/[category]/img_NNN.jpg
"""
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import os
import json
import base64
import re

BASE = r"C:\Users\lianxiang\Desktop\split_by_style\split_by_style\6大类"
OUT_DIR = r"C:\Users\lianxiang\WorkBuddy\2026-07-23-09-09-54\style-gallery"
IMG_DIR = os.path.join(OUT_DIR, "images")
DATA_DIR = os.path.join(OUT_DIR, "data")

os.makedirs(IMG_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

XDR_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
A_NS = 'http://schemas.openxmlformats.org/drawingml/2006/main'
R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

ns_map = {
    'xdr': XDR_NS,
    'a': A_NS,
    'r': R_NS,
}

CATEGORY_MAP = {
    "01_动漫影视.xlsx": "01_anime",
    "02_科幻游戏.xlsx": "02_scifi",
    "03_绘画艺术.xlsx": "03_art",
    "04_商业设计.xlsx": "04_design",
    "05_写实时尚.xlsx": "05_fashion",
    "06_其他综合.xlsx": "06_other",
}

def extract_image_row_map(xlsx_path):
    """Parse drawing1.xml to get {row_index: (rId, image_filename)} mapping.
    row_index is 0-based in the drawing, so sheet row = row_index + 1
    """
    mapping = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # Check if drawing exists
        drawing_names = [n for n in z.namelist() if 'drawing' in n and n.endswith('.xml')]
        if not drawing_names:
            return mapping
        
        drawing_xml = z.read(drawing_names[0]).decode('utf-8')
        
        # Find rels
        rels_names = [n for n in z.namelist() if 'drawing' in n and n.endswith('.rels')]
        if not rels_names:
            return mapping
        
        rels_xml = z.read(rels_names[0]).decode('utf-8')
        
        # Parse rels: rId -> image filename
        rels_root = ET.fromstring(rels_xml)
        rid_to_file = {}
        for rel in rels_root:
            rid = rel.get('Id')
            target = rel.get('Target')
            if target:
                # Target is like "../media/image1.jpeg"
                fname = os.path.basename(target)
                rid_to_file[rid] = fname
        
        # Parse drawing to get row -> rId mapping
        root = ET.fromstring(drawing_xml)
        
        for anchor in root:
            tag = anchor.tag.split('}')[-1] if '}' in anchor.tag else anchor.tag
            
            if tag not in ('oneCellAnchor', 'twoCellAnchor'):
                continue
            
            from_el = anchor.find(f'{{{XDR_NS}}}from')
            if from_el is None:
                continue
            
            col_el = from_el.find(f'{{{XDR_NS}}}col')
            row_el = from_el.find(f'{{{XDR_NS}}}row')
            
            if col_el is None or row_el is None:
                continue
            
            col = int(col_el.text)
            row = int(row_el.text)
            
            # Find the picture's embed rId
            pic = anchor.find(f'{{{XDR_NS}}}pic')
            if pic is None:
                continue
            
            blip = pic.find(f'.//{{{A_NS}}}blip')
            if blip is None:
                continue
            
            embed = blip.get(f'{{{R_NS}}}embed')
            if embed and embed in rid_to_file:
                # row is 0-based; sheet row = row + 1 (since row 0 = header row 1)
                sheet_row = row + 1
                mapping[sheet_row] = (embed, rid_to_file[embed], col)
    
    return mapping

def extract_images(xlsx_path, category, row_map):
    """Extract image files from xlsx and save to disk. Return {row: saved_filename}."""
    saved = {}
    cat_img_dir = os.path.join(IMG_DIR, category)
    os.makedirs(cat_img_dir, exist_ok=True)
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        media_files = {}
        for name in z.namelist():
            if name.startswith('xl/media/') and name.endswith('.jpeg'):
                fname = os.path.basename(name)
                media_files[fname] = name
        
        for sheet_row, (rid, img_fname, col) in row_map.items():
            if img_fname in media_files:
                zname = media_files[img_fname]
                data = z.read(zname)
                # Save with a clean name
                out_name = f"img_{sheet_row:05d}.jpg"
                out_path = os.path.join(cat_img_dir, out_name)
                with open(out_path, 'wb') as f:
                    f.write(data)
                saved[sheet_row] = f"images/{category}/{out_name}"
    
    return saved

def process_file(fname, category):
    fpath = os.path.join(BASE, fname)
    print(f"\nProcessing: {fname} -> {category}")
    
    # Step 1: Get image row mapping
    row_map = extract_image_row_map(fpath)
    print(f"  Found {len(row_map)} image placements")
    
    # Step 2: Extract images
    saved_images = extract_images(fpath, category, row_map)
    print(f"  Extracted {len(saved_images)} images")
    
    # Step 3: Read Excel data
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    
    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True), start=2):
        title = str(row[0]).strip() if row[0] else ""
        prompt = str(row[1]).strip() if row[1] else ""
        
        if not title and not prompt:
            continue
        
        # Truncate very long prompts for JSON storage
        if len(prompt) > 8000:
            prompt = prompt[:8000] + "\n...[截断]"
        
        img_path = saved_images.get(row_idx, "")
        
        entries.append({
            "id": row_idx - 1,
            "title": title,
            "prompt": prompt,
            "image": img_path,
        })
    
    wb.close()
    
    print(f"  Total entries: {len(entries)}")
    
    # Save JSON
    json_path = os.path.join(DATA_DIR, f"{category}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(entries, f, ensure_ascii=False, separators=(',', ':'))
    
    print(f"  Saved JSON: {json_path}")
    
    # Stats
    with_img = sum(1 for e in entries if e["image"])
    print(f"  Entries with images: {with_img}")
    
    return len(entries), with_img

# Process all files
total_entries = 0
total_images = 0
for fname, category in CATEGORY_MAP.items():
    entries_count, img_count = process_file(fname, category)
    total_entries += entries_count
    total_images += img_count

print(f"\n{'='*60}")
print(f"DONE! Total: {total_entries} entries, {total_images} images")
