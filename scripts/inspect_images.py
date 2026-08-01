import openpyxl
import os
import json

base = r"C:\Users\lianxiang\Desktop\split_by_style\split_by_style\6大类"
files = sorted(os.listdir(base))

# Pick the smallest file first to understand structure
fpath = os.path.join(base, "06_其他综合.xlsx")
wb = openpyxl.load_workbook(fpath)
ws = wb.active

# Check column headers
print("=== Headers ===")
for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
    for cell in row:
        print(f"  Col {cell.column}: '{cell.value}' (col_letter={cell.column_letter})")

# Check what's in the image column - look at first 5 rows
print("\n=== Image column (Col D) first 5 rows ===")
for row in ws.iter_rows(min_row=2, max_row=6, min_col=4, max_col=4, values_only=False):
    for cell in row:
        print(f"  Row {cell.row}: value='{cell.value}' type={type(cell.value).__name__}")

# Also check col C (空列)
print("\n=== Col C first 5 rows ===")
for row in ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=3, values_only=False):
    for cell in row:
        print(f"  Row {cell.row}: value='{cell.value}' type={type(cell.value).__name__}")

wb.close()

# Now check if images are embedded in the xlsx as drawing/xml
print("\n=== Checking xlsx internals ===")
import zipfile
with zipfile.ZipFile(fpath, 'r') as z:
    for name in z.namelist():
        if 'image' in name.lower() or 'media' in name.lower() or 'drawing' in name.lower():
            info = z.getinfo(name)
            print(f"  {name} ({info.file_size} bytes)")

