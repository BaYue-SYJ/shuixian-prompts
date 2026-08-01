import openpyxl
import os
import random

base = r"C:\Users\lianxiang\Desktop\split_by_style\split_by_style\6大类"
files = sorted(os.listdir(base))

for fname in files:
    if not fname.endswith('.xlsx'):
        continue
    fpath = os.path.join(base, fname)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        
        # Get all titles
        titles = []
        sample_prompts = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True)):
            if row[0]:
                titles.append(str(row[0]))
            if row[1] and len(sample_prompts) < 3:
                sample_prompts.append(str(row[1])[:200])
        
        print(f"Total data rows: {len(titles)}")
        print(f"\n--- All Titles (first 80) ---")
        for t in titles[:80]:
            print(f"  - {t}")
        
        if len(titles) > 80:
            print(f"  ... and {len(titles)-80} more titles")
        
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
