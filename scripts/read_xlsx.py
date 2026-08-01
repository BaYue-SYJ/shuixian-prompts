import openpyxl
import os

base = r"C:\Users\lianxiang\Desktop\split_by_style\split_by_style\6大类"
files = sorted(os.listdir(base))

for fname in files:
    if not fname.endswith('.xlsx'):
        continue
    fpath = os.path.join(base, fname)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        print(f"Sheet name: {ws.title}")

        # Get headers from first row
        rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
        if rows:
            print(f"Headers: {rows[0]}")
            print(f"Total rows (approx): {ws.max_row}")
            print(f"\n--- Sample rows ---")
            for i, row in enumerate(rows[1:6], 1):
                # Print each cell, truncated
                for j, cell in enumerate(row):
                    if cell is not None:
                        cell_str = str(cell)[:500]
                        print(f"  Row{i} Col{j+1}: {cell_str}")
                    else:
                        print(f"  Row{i} Col{j+1}: [None]")
                print()
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
